"""
Section-by-section AI extraction of IP-analysis fields from an Instagram video.

Pipeline (per post):
    1. Upload the local video file to Gemini once (cached by sha256(path)).
    2. For each logical section (post_info, context, character, plot, drawing,
       messaging, comments), build a prompt from `ip_fields.xlsx` rows and call
       `gemini_client.models.generate_content` with `response_mime_type="application/json"`
       and `response_schema=<SectionModel>`. Gemini enforces shape + closed enums.
    3. Merge per-section outputs with meta-derived fields (timestamp, durations,
       counts) pulled from the Document, then normalize through schema_tools'
       `InstagramPostAnalysis` schema for typed output.

Question definitions live in `src/scripts/ips/resources/ip_fields.xlsx`. The
section→question grouping and question→json-field mapping are owned by this
module (not by the xlsx category column) so the xlsx remains a human-facing
reference and code stays stable when the xlsx is reorganised.
"""
from __future__ import annotations

import hashlib
import logging
import os
import re
import time
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, Tuple

import openpyxl
from pydantic import BaseModel, Field

from google.genai import types as genai_types

from schema_tools import normalize_record

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

GEMINI_MODEL = "gemini-2.5-pro"
DEFAULT_XLSX = Path(__file__).parent / "resources" / "ip_fields.xlsx"

# Maps each logical section (matches the schema_tools composite naming) to the
# question numbers from ip_fields.xlsx that should be answered by the AI in
# that section's prompt. Meta-derived non-AI rows (Q1, Q5, Q7, Q10, Q11, Q13,
# Q68) and out-of-scope rows (Q3, Q4, Q44, Q72 — `field=None` in the xlsx) are
# excluded by design.
SECTIONS: Dict[str, List[int]] = {
    "post_info": [2, 6, 8, 9, 12, 14],
    "context":   [15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28],
    "character": [29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 46],
    "plot":      [40, 41, 42, 43, 45],
    "drawing":   [47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60],
    "messaging": [61, 62, 63, 64, 65, 66, 67],
    "comments":  [69, 70, 71],
}

# Maps a question number to the JSON field name(s) the model should fill.
# `characters[].<name>` indicates a per-character field inside the character
# section's `characters` list. Composite questions (Q45) list multiple fields.
FIELD_FOR_Q: Dict[int, str] = {
    # Post Info
    2:  "caption_languages",
    6:  "audio_class",
    8:  "has_cta",
    9:  "ctas",
    12: "series_or_linked",
    14: "loopability",
    # Context
    15: "temporal_context",
    16: "general_post_type",
    17: "objects",
    18: "pov",
    19: "setting",
    20: "theme",
    21: "following_trend",
    22: "trend_name",
    23: "emotional_tone",
    24: "primary_purpose",
    25: "relatability_score",
    26: "target_gender",
    27: "target_age",
    28: "target_interests",
    # Character section (mixed: top-level + per-character)
    29: "has_characters",
    30: "character_count_label",
    31: "characters[].outlook",
    32: "characters[].personality",
    33: "characters[].emotions",
    34: "characters[].key_descriptives",
    35: "characters[].archetypes",
    36: "characters[].facial_expression_changes",
    37: "characters[].visual_appeal",
    38: "characters[].gaze_direction",
    39: "has_voiceover",
    46: "characters[].outlined",
    # Plot
    40: "summary",
    41: "realistic",
    42: "interactions",
    43: "scene_changes",
    45: "pacing_cuts, pacing_label",
    # Drawing
    47: "outline_kind",
    48: "colour_mode",
    49: "lighting",
    50: "colour_palette",
    51: "colour_count_bucket",
    52: "pure_background",
    53: "pure_bg_colour",
    54: "bg_description",
    55: "graphic_elements",
    56: "animation_style",
    57: "frame_rate_feel",
    58: "dynamic_effects",
    59: "aesthetic_traits",
    60: "texture",
    # Messaging
    61: "has_text_overlays",
    62: "text_overlay_categories",
    63: "text_overlays_verbatim",
    64: "typography",
    65: "has_supportive_message",
    66: "message_summary",
    67: "encourages_action",
    # Comments
    69: "top_liked_comment",
    70: "top_keywords",
    71: "top_emojis",
}

# Apify Instagram post types → schema_tools `post_format` enum values.
APIFY_POST_TYPE_TO_FORMAT: Dict[str, str] = {
    "Image":   "Static Image",
    "Sidecar": "Image Carousel",
    "Video":   "Single Video",
    "Reel":    "Single Video",
    "Story":   "Single Video",
}

# ---------------------------------------------------------------------------
# Pydantic response models — one per Gemini call.
# These mirror the schema_tools composite types field-for-field. Closed
# Catalogue/Multi-catalogue use Literal[...] (Gemini honors as enum). Extensible
# types stay plain str / list[str] with prompt-level guidance.
# ---------------------------------------------------------------------------


class PostInfoSectionOut(BaseModel):
    caption_languages: List[str] = Field(default_factory=list)
    audio_class: str = ""
    has_cta: bool = False
    ctas: List[str] = Field(default_factory=list)
    series_or_linked: Literal["Yes", "No", "Unclear"] = "Unclear"
    loopability: Literal["Yes", "No", "Partial"] = "No"


class ContextSectionOut(BaseModel):
    temporal_context: str = ""
    general_post_type: str = ""
    objects: List[str] = Field(default_factory=list)
    pov: Literal["First-person", "Third-person", "Mixed", "Other"] = "Other"
    setting: Literal[
        "Indoor", "Outdoor", "Both/Mixed", "Studio/Plain background", "Indeterminate"
    ] = "Indeterminate"
    theme: List[str] = Field(default_factory=list)
    following_trend: bool = False
    trend_name: Optional[str] = None
    emotional_tone: List[str] = Field(default_factory=list)
    primary_purpose: str = ""
    relatability_score: int = 0
    target_gender: List[Literal["F", "M", "Neutral"]] = Field(default_factory=list)
    target_age: List[Literal["Kids", "Teens", "Young Adults", "Adults", "All"]] = Field(default_factory=list)
    target_interests: List[Literal["Fandom", "Specific interest", "General"]] = Field(default_factory=list)


class CharacterOut(BaseModel):
    name: str = ""
    outlook: str = ""
    personality: str = ""
    emotions: str = ""
    key_descriptives: List[str] = Field(default_factory=list)
    archetypes: List[str] = Field(default_factory=list)
    facial_expression_changes: int = 0
    visual_appeal: Literal[
        "Chibi", "Minimalist", "Hyper-detailed", "Realistic cartoon", "Stylized", "Other"
    ] = "Other"
    gaze_direction: Literal["Yes", "No", "Occasional"] = "No"
    outlined: Literal["Outlined", "Not Outlined", "Partial"] = "Not Outlined"


class CharacterSectionOut(BaseModel):
    has_characters: bool = False
    character_count_label: Literal[
        "1", "2", "3", "4", "5+", "Group scene", "Uncountable", "None"
    ] = "None"
    has_voiceover: Literal["Yes", "No", "Text-to-speech", "Music only"] = "No"
    characters: List[CharacterOut] = Field(default_factory=list)


class PlotSectionOut(BaseModel):
    summary: str = ""
    realistic: Literal["Yes", "No", "Yes but has fantastic elements"] = "Yes"
    interactions: Optional[str] = None
    scene_changes: int = 0
    pacing_cuts: int = 0
    pacing_label: Literal["Fast", "Medium", "Slow"] = "Medium"


class DrawingSectionOut(BaseModel):
    outline_kind: Optional[str] = None
    colour_mode: Literal["Bright Mode", "Dark Mode"] = "Bright Mode"
    lighting: Literal["Bright/High-Key", "Moody/Low-Key", "High Contrast", "Flat"] = "Flat"
    colour_palette: Literal["Monochrome", "Pastel", "Saturated", "Earthy", "Neon"] = "Pastel"
    colour_count_bucket: Literal["<7", "7-20", "21 or more"] = "7-20"
    pure_background: bool = False
    pure_bg_colour: Optional[str] = None
    bg_description: Optional[str] = None
    graphic_elements: List[str] = Field(default_factory=list)
    animation_style: str = ""
    frame_rate_feel: Literal["Fluid (24+ fps)", "Classic (12-24)", "Limited/choppy (<12)", "Other"] = "Other"
    dynamic_effects: List[str] = Field(default_factory=list)
    aesthetic_traits: List[str] = Field(default_factory=list)
    texture: Literal["Clean", "Noisy/Grainy", "Textured", "Watercolor"] = "Clean"


class MessagingSectionOut(BaseModel):
    has_text_overlays: bool = False
    text_overlay_categories: List[str] = Field(default_factory=list)
    text_overlays_verbatim: Optional[str] = None
    typography: Optional[str] = None
    has_supportive_message: bool = False
    message_summary: Optional[str] = None
    encourages_action: bool = False


class CommentInsightsSectionOut(BaseModel):
    top_liked_comment: Optional[str] = None
    top_keywords: List[str] = Field(default_factory=list)
    top_emojis: List[str] = Field(default_factory=list)


SECTION_MODEL: Dict[str, type[BaseModel]] = {
    "post_info": PostInfoSectionOut,
    "context":   ContextSectionOut,
    "character": CharacterSectionOut,
    "plot":      PlotSectionOut,
    "drawing":   DrawingSectionOut,
    "messaging": MessagingSectionOut,
    "comments":  CommentInsightsSectionOut,
}

SECTION_HEADER: Dict[str, str] = {
    "post_info": "Post Info",
    "context":   "Context",
    "character": "Characters",
    "plot":      "Plot",
    "drawing":   "Drawing / visual style",
    "messaging": "Messaging / text overlays",
    "comments":  "Comments insights",
}

# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------


def load_questions(xlsx_path: Path = DEFAULT_XLSX) -> Dict[int, Dict[str, Any]]:
    """Read the Fields sheet and return {question_no: row_dict}."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Fields"]
    rows: Dict[int, Dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        no, category, output, datatype, question, responses, field, notes, claude_notes = row
        if no is None:
            continue
        try:
            no_int = int(no)
        except (TypeError, ValueError):
            continue
        rows[no_int] = {
            "no": no_int,
            "category": category,
            "output": output,
            "datatype": datatype,
            "question": question,
            "responses": responses,
            "field": field,
            "notes": notes,
            "claude_notes": claude_notes,
        }
    return rows


# ---------------------------------------------------------------------------
# Prompt building
# ---------------------------------------------------------------------------


_EXTENSIBLE_TYPES = {"Extensible Catalogue", "Extensible Multi-catalogue"}
_CLOSED_TYPES = {"Catalogue", "Multi-catalogue", "Boolean"}


def _options_hint(datatype: Optional[str]) -> str:
    if datatype in _EXTENSIBLE_TYPES:
        return "Use one of these options when appropriate, or propose a new value if none fits."
    if datatype in _CLOSED_TYPES:
        return "Answer must be one of the listed values exactly."
    if datatype and datatype.startswith("List["):
        return "Answer is a list (one entry per character / item). Allowed values: see schema."
    return ""


def build_section_prompt(
    section: str,
    questions: Dict[int, Dict[str, Any]],
    caption: str,
) -> str:
    """Build the Gemini prompt for one section."""
    q_nos = SECTIONS[section]
    header = SECTION_HEADER[section]
    lines: List[str] = [
        f"You are analyzing an Instagram video. Answer the questions below about *{header}*.",
        "",
        "Return a single JSON object that exactly matches the response schema.",
        "Do not invent fields and do not wrap the JSON in markdown.",
        "",
        f"Caption: {caption or '(empty)'}",
        "",
        "Questions:",
    ]
    for i, no in enumerate(q_nos, start=1):
        q = questions.get(no)
        if q is None:
            continue
        target = FIELD_FOR_Q.get(no, "?")
        responses = q.get("responses") or ""
        datatype = q.get("datatype") or ""
        hint = _options_hint(datatype)
        line = f"{i}. ({target}) {q['question']}"
        if responses:
            line += f"\n   Responses: {responses}"
        if hint:
            line += f"\n   {hint}"
        lines.append(line)

    if section == "character":
        lines += [
            "",
            "Character section guidance:",
            "- Set `has_characters` and `character_count_label` at the top level.",
            "- For each detected character, append one entry to `characters` with all per-character fields.",
            "- When `has_characters` is false, return an empty `characters` list and set `character_count_label` to 'None'.",
            "- `has_voiceover` is post-level (audio track), not per-character.",
        ]
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Gemini upload + per-section call
# ---------------------------------------------------------------------------


def _sha256_file(path: str | Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _wait_for_active(client, file_obj, timeout_s: int = 300, poll_s: float = 2.0):
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        f = client.files.get(name=file_obj.name)
        if f.state.name == "ACTIVE":
            return f
        if f.state.name == "FAILED":
            raise RuntimeError(f"Gemini file processing failed: {f.name}")
        time.sleep(poll_s)
    raise TimeoutError(f"Gemini file {file_obj.name} did not become ACTIVE within {timeout_s}s")


_UPLOAD_CACHE: Dict[str, Any] = {}


def upload_video_cached(client, video_path: str) -> Any:
    """Upload `video_path` to Gemini Files exactly once per process per content hash."""
    key = _sha256_file(video_path)
    cached = _UPLOAD_CACHE.get(key)
    if cached is not None:
        try:
            refreshed = client.files.get(name=cached.name)
            if refreshed.state.name == "ACTIVE":
                log.info("[gemini-upload] cache hit %s -> %s", video_path, cached.name)
                return refreshed
        except Exception:  # file expired / deleted upstream — fall through to re-upload
            pass

    log.info("[gemini-upload] uploading %s", video_path)
    uploaded = client.files.upload(file=video_path)
    uploaded = _wait_for_active(client, uploaded)
    _UPLOAD_CACHE[key] = uploaded
    return uploaded


def describe_section(
    client,
    uploaded_file,
    section: str,
    questions: Dict[int, Dict[str, Any]],
    caption: str,
    *,
    temperature: float = 0.3,
) -> Dict[str, Any]:
    """One Gemini call for one section. Returns the parsed dict (model_dump)."""
    prompt = build_section_prompt(section, questions, caption)
    model_cls = SECTION_MODEL[section]
    resp = client.models.generate_content(
        model=GEMINI_MODEL,
        contents=[uploaded_file, prompt],
        config=genai_types.GenerateContentConfig(
            temperature=temperature,
            response_mime_type="application/json",
            response_schema=model_cls,
        ),
    )
    parsed = resp.parsed
    if parsed is None:
        # SDK couldn't auto-parse; fall back to raw text → model_validate_json
        parsed = model_cls.model_validate_json(resp.text)
    return parsed.model_dump()


def describe_video_sectioned(
    client,
    video_path: str,
    caption: str,
    *,
    sections: Optional[List[str]] = None,
    questions: Optional[Dict[int, Dict[str, Any]]] = None,
) -> Dict[str, Dict[str, Any]]:
    """Upload once, then call once per section. Per-section failures are isolated."""
    sections = sections or list(SECTIONS.keys())
    questions = questions or load_questions()
    uploaded = upload_video_cached(client, video_path)

    results: Dict[str, Dict[str, Any]] = {}
    for section in sections:
        try:
            log.info("[gemini-section] %s", section)
            results[section] = describe_section(client, uploaded, section, questions, caption)
        except Exception as exc:  # noqa: BLE001
            log.exception("[gemini-section] %s failed: %s", section, exc)
            results[section] = {"_error": str(exc)}
    return results


# ---------------------------------------------------------------------------
# Meta-derived field helpers
# ---------------------------------------------------------------------------


_HASHTAG_RE = re.compile(r"(?<!\w)#\w+", re.UNICODE)
_MENTION_RE = re.compile(r"(?<!\w)@\w[\w.]*", re.UNICODE)
# Coarse emoji range: BMP symbols + supplementary planes commonly used by Apify.
_EMOJI_RE = re.compile(
    "["
    "\U0001F300-\U0001FAFF"
    "\U00002600-\U000027BF"
    "\U0001F000-\U0001F1FF"
    "]",
    flags=re.UNICODE,
)


def _word_count(text: str) -> int:
    return len([w for w in re.split(r"\s+", text.strip()) if w])


def _emoji_pct(text: str, word_count: int) -> float:
    if not word_count:
        return 0.0
    return round(100.0 * len(_EMOJI_RE.findall(text)) / word_count, 2)


def fill_meta_fields(doc_data: Dict[str, Any], ai_results: Dict[str, Dict[str, Any]]) -> None:
    """Mutate `ai_results` in place: add meta-derived fields to the right sections."""
    body = doc_data.get("body") or ""
    word_count = _word_count(body)

    post_info = ai_results.setdefault("post_info", {})
    post_info["timestamp"] = doc_data.get("timestamp")
    post_info["video_duration_s"] = doc_data.get("video_duration") or doc_data.get("video_duration_s")
    post_info["post_type"] = doc_data.get("post_type")
    post_info["post_format"] = APIFY_POST_TYPE_TO_FORMAT.get(doc_data.get("post_type") or "")
    post_info["hashtag_count"] = len(_HASHTAG_RE.findall(body))
    post_info["mention_count"] = len(_MENTION_RE.findall(body))
    post_info["caption_word_count"] = word_count
    post_info["emoji_pct"] = _emoji_pct(body, word_count)
    post_info["likes"] = doc_data.get("likes")
    post_info["views"] = doc_data.get("views")
    post_info["shares"] = doc_data.get("shares")

    comments = ai_results.setdefault("comments", {})
    comments["comment_count"] = doc_data.get("n_comments")


# ---------------------------------------------------------------------------
# Top-level record assembly + schema_tools normalization
# ---------------------------------------------------------------------------


def _split_character_section(char_section: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """Split CharacterSectionOut output into (top_level_fields, characters_list)."""
    if "_error" in char_section:
        return {}, []
    top = {
        "character_count_label": char_section.get("character_count_label"),
        "has_voiceover": char_section.get("has_voiceover"),
    }
    chars = char_section.get("characters") or []
    return top, list(chars)


def build_record(doc_data: Dict[str, Any], ai_results: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    """Merge AI + meta into the InstagramPostAnalysis shape and normalize."""
    fill_meta_fields(doc_data, ai_results)
    char_top, chars = _split_character_section(ai_results.get("character", {}))

    record = {
        "post_url": doc_data.get("url"),
        "character_count_label": char_top.get("character_count_label"),
        "has_voiceover": char_top.get("has_voiceover"),
        "post_info": ai_results.get("post_info", {}),
        "context": ai_results.get("context", {}),
        "characters": chars,
        "plot": ai_results.get("plot", {}),
        "drawing": ai_results.get("drawing", {}),
        "messaging": ai_results.get("messaging", {}),
        "comment_insights": ai_results.get("comments", {}),
    }
    return normalize_record(record, "InstagramPostAnalysis")


# ---------------------------------------------------------------------------
# High-level convenience
# ---------------------------------------------------------------------------


def analyse_post(
    client,
    doc_data: Dict[str, Any],
    *,
    sections: Optional[List[str]] = None,
    questions: Optional[Dict[int, Dict[str, Any]]] = None,
) -> Optional[Dict[str, Any]]:
    """Run the full pipeline for one post. Returns None when no video is available."""
    video_path = doc_data.get("video_filename")
    if not video_path or not os.path.exists(video_path):
        log.warning("[analyse_post] skipping %s: no video at %s", doc_data.get("url"), video_path)
        return None
    caption = doc_data.get("body") or ""
    ai_results = describe_video_sectioned(
        client, video_path, caption, sections=sections, questions=questions
    )
    return build_record(doc_data, ai_results)
